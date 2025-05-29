#%% md
# ### Imports
# 
#%%
import traceback
from datetime import datetime

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from sklearn.metrics import confusion_matrix, classification_report
from sklearn.model_selection import train_test_split
from tabulate import tabulate
from tensorflow.keras import Model, Input
from tensorflow.keras.layers import Conv2D, MaxPooling2D, Dropout, Flatten, Dense
from tensorflow.keras.optimizers import Adam
from tensorflow.keras.regularizers import l1
#%% md
# ### Load Feature Matrices
#%%
def read_csv_to_4d_np(name):
    """ Reads the csv file and changes it to 4D numpy array """

    print('{} : Reading 4D numpy {}'.format(datetime.now(), name))
    reshaped = []
    try:
        np_data = np.loadtxt('{}.csv'.format(name), delimiter=',')
        print('{} : Shape {}'.format(datetime.now(), np_data.shape))
        reshaped = np_data.reshape(np_data.shape[1], 4, 62, np_data.shape[0] // 248)
        # Transpose to reorder dimensions from (None, 4, 62, Feature) to (None, 62, Features, 4) for use in CNN
        transposed = reshaped.transpose(0, 2, 3, 1)
        print('{} : Read 4D numpy {} : Shape {}'.format(datetime.now(), name, transposed.shape))
    except Exception as e:
        print('{} : Error reading 4D numpy: {}'.format(datetime.now(), e))
        traceback.print_exc()
    return transposed
#%%
pcc_matrix_8s = read_csv_to_4d_np('pcc_matrix_8s')
pcc_matrix_12s = read_csv_to_4d_np('pcc_matrix_12s')

conn_matrix_8s = read_csv_to_4d_np('conn_matrix_8s')
conn_matrix_12s = read_csv_to_4d_np('conn_matrix_12s')

pca_matrix_8s = read_csv_to_4d_np('pca_matrix_8s')
pca_matrix_12s = read_csv_to_4d_np('pca_matrix_12s')
sc_matrix_8s = read_csv_to_4d_np('sc_matrix_8s')
sc_matrix_12s = read_csv_to_4d_np('sc_matrix_12s')
#%% md
# ### Data Labels
#%%
# Since the labels are constant for trials of all participants, I can just copy them from the readme.txt.

# I would also have to map the labels from -1, 0, 1 to 0, 1, 2 to fit in the NN models optimizers.

labels_array = [1, 0, -1, -1, 0, 1, -1, 0, 1, 1, 0, -1, 0, 1, -1]

def get_label_from_feature_set(feature_set, labels):
    print('{} : Creating Labels Array'.format(datetime.now()))
    no_participants = 15
    no_labels = 15
    no_segments = feature_set.shape[0] // (no_participants * no_labels)

    # Repeat each label by the number of segments
    repeated_labels = np.repeat(labels, no_segments)

    # Repeat the entire sequence by the number of participants
    label_array = np.tile(repeated_labels, no_participants)

    # Map the labels ( -1 => 0, 0 => 1, 1 => 2 )
    mapped_labels = np.where(label_array == -1, 0,
                             np.where(label_array == 0, 1,
                                      np.where(label_array == 1, 2, -1)))

    print('{} : Created Labels Array : Shape {}'.format(datetime.now(), mapped_labels.shape))
    return mapped_labels
#%%
labels_8s = get_label_from_feature_set(pcc_matrix_8s, labels_array)
np.savetxt(f'labels_8s.csv', labels_8s, delimiter=',')
labels_12s = get_label_from_feature_set(pcc_matrix_12s, labels_array)
np.savetxt(f'labels_12s.csv', labels_12s, delimiter=',')
#%% md
# ### CNN Model
#%%
# CNN Model
class CNN:
    def __init__(self, input_shape):
        self.input_shape = input_shape
        self.model = None

    def build(self):
        """
        Build the CNN model for feature extraction.
        """
        input_layer = Input(shape=self.input_shape)
        x = Conv2D(32, (3, 3), activation='relu', padding='same')(input_layer)
        x = MaxPooling2D((2, 2))(x)
        x = Conv2D(64, (3, 3), activation='relu', padding='same')(x)
        x = MaxPooling2D((2, 2))(x)
        x = Flatten()(x)  # Flatten the feature map for the next stage
        self.model = Model(inputs=input_layer, outputs=x)
        return self.model

#%% md
# ### SAE Model
#%%
class SparseAutoencoder:
    def __init__(self, input_dim, encoding_dim, sparsity=1e-5):
        self.input_dim = input_dim
        self.encoding_dim = encoding_dim
        self.sparsity = sparsity
        self.encoder = None

    def build(self, input_layer):
        """
        Build the sparse autoencoder layers.
        """
        encoded = Dense(self.encoding_dim, activation='relu', activity_regularizer=l1(self.sparsity))(input_layer)
        return encoded

#%% md
# ### DNN Model
#%%
class DNN:
    def __init__(self, num_classes):
        self.num_classes = num_classes

    def build(self, input_layer):
        """
        Build the DNN layers for classification.
        """
        x = Dense(128, activation='relu')(input_layer)
        x = Dropout(0.5)(x)  # Dropout for regularization
        output_layer = Dense(self.num_classes, activation='softmax')(x)
        return output_layer

#%% md
# # Hybrid Model
#%%
class HybridNeuralNetwork:
    def __init__(self, input_shape, encoding_dim, num_classes, sparsity=1e-5, learning_rate=0.001, epochs=50):
        self.input_shape = input_shape
        self.encoding_dim = encoding_dim
        self.num_classes = num_classes
        self.sparsity = sparsity
        self.model = None
        self.learning_rate = learning_rate
        self.epochs = epochs

    def build_model(self):
        """
        Build the hybrid neural network by combining CNN, SAE, and DNN.
        """
        # Build CNN
        cnn = CNN(self.input_shape)
        cnn_model = cnn.build()

        # Build Sparse Autoencoder
        sae = SparseAutoencoder(input_dim=cnn_model.output_shape[1], encoding_dim=self.encoding_dim,
                                sparsity=self.sparsity)
        encoded = sae.build(cnn_model.output)

        # Build DNN
        dnn = DNN(self.num_classes)
        output_layer = dnn.build(encoded)

        # Combine all components into a single model
        self.model = Model(inputs=cnn_model.input, outputs=output_layer)

    def compile_model(self):
        """
        Compile the model with a specified learning rate.
        """
        if self.model is None:
            raise ValueError("Model has not been built yet. Call build_model() first.")
        self.model.compile(optimizer=Adam(learning_rate=self.learning_rate),
                           loss='sparse_categorical_crossentropy',
                           metrics=['accuracy'])

    def train_model(self, X_train, y_train, X_val, y_val, batch_size=32):
        """
        Train the model on the given training data.
        """
        if self.model is None:
            raise ValueError("Model has not been built yet. Call build_model() first.")
        history = self.model.fit(X_train, y_train,
                                 validation_data=(X_val, y_val),
                                 epochs=self.epochs,
                                 batch_size=batch_size)
        return history

    def evaluate_model(self, X_test, y_test):
        """
        Evaluate the model on the test data.
        """
        if self.model is None:
            raise ValueError("Model has not been built yet. Call build_model() first.")
        return self.model.evaluate(X_test, y_test)

    def predict(self, X):
        """
        Make predictions using the trained model.
        """
        if self.model is None:
            raise ValueError("Model has not been built yet. Call build_model() first.")
        return self.model.predict(X)

#%% md
# ### Plotting
#%%
def plot_history(history, history_type, feature_type):
    plt.plot(history.history[feature_type.split('-')[0]], label=f'{history_type} {feature_type}')
    plt.title(f'{history_type} {feature_type}')
    plt.xlabel('Epochs')
    plt.ylabel(feature_type.split('-')[0])
    plt.legend()
    plt.show()
#%% md
# ### Data split
#%%
# A function to split data into 80-20 training test
def split_data(features, labels, test_size=0.2):
    print('{} : Splitting Data and Labels'.format(datetime.now()))
    x_train, x_test, y_train, y_test = train_test_split(features, labels, test_size=test_size, random_state=42)
    print(f"Training set shape: {x_train.shape}, Training labels shape: {y_train.shape}")
    print(f"Testing set shape: {x_test.shape}, Testing labels shape: {y_test.shape}")
    print('{} : Splitting Data and Labels, DONE'.format(datetime.now()))
    return x_train, x_test, y_train, y_test
#%% md
# ### History Printing
#%%
def print_history_average(_history, dict, feature_type):
    # Calculate final averages
    final_avg = {}
    for key in _history.history.keys():
        final_avg[key] = np.mean(_history.history[key])

    # Print final averages
    print("\nFinal Averages:")
    for key, value in final_avg.items():
        print(f"{key} (average): {value:.4f}")

    dict[feature_type] = final_avg
    return dict
#%% md
# ### Parameter and Results Printing and Saving
#%%
def save_params_and_results(model_dict, results_dict, cm_dict, report_dict, learning_rate, epochs, file_path, instance):
    # Prepare data for tabulation
    param_rows = []
    for model_name, model in model_dict.items():
        for layer in model.model.layers:
            # Access all weights and biases for layers with parameters
            weights = layer.get_weights()

            # If layer has weights (like Dense, Conv2D)
            if weights:
                for i, weight_array in enumerate(weights):
                    if i == 0:
                        # First array in weights is the kernel (weights)
                        param_name = f"{layer.name}_weights"
                    elif i == 1:
                        # Second array in weights is the bias
                        param_name = f"{layer.name}_bias"

                    param_rows.append([model_name, param_name, str(weight_array.shape), weight_array])

            # Access regularizer parameters (if any)
            if hasattr(layer, 'kernel_regularizer') and layer.kernel_regularizer is not None:
                regularizer = layer.kernel_regularizer
                param_rows.append([model_name, f"{layer.name}_regularizer", str(regularizer), 'N/A'])

    # Headers for the table
    param_headers = ['Model', 'Parameter Name', 'Shape', 'Values (First Few)']

    # Function to truncate large weight arrays for display
    def truncate_values(values, max_elements=10):
        """Truncate the displayed parameter values to show only the first few elements for readability."""
        if values.size > max_elements:
            return str(values.flatten()[:max_elements]) + '...'
        return str(values.flatten())

    # Modify rows to truncate values for readability
    rows_with_truncated_values = [
        [row[0], row[1], row[2], str(truncate_values(row[3]))]
        for row in param_rows
    ]
    # Display the table
    print(tabulate(rows_with_truncated_values, headers=param_headers, tablefmt='grid'))

    result_rows = []
    feature_set_names = sorted(set(key.split('_test')[0] for key in results_dict.keys()))  # Extract unique model names

    for model in feature_set_names:
        train_metrics = results_dict.get(model, {})
        test_metrics = results_dict.get(model + '_test', {})

        # Create a row for each model, combining training and test metrics
        row = [
            model,
            train_metrics.get('accuracy', ''),
            train_metrics.get('loss', ''),
            train_metrics.get('val_accuracy', ''),
            train_metrics.get('val_loss', ''),
            test_metrics.get('accuracy', ''),
            test_metrics.get('loss', '')
        ]
        result_rows.append(row)

    # Headers for the table
    result_headers = ['Model', 'Train Accuracy', 'Train Loss', 'Validation Accuracy', 'Validation Loss',
                      'Test Accuracy', 'Test Loss']

    # Display table using tabulate
    print(tabulate(result_rows, headers=result_headers, tablefmt='grid'))

    # Ensure param_rows and result_rows are 2D
    param_rows = [list(row) for row in param_rows]
    result_rows = [list(row) for row in result_rows]

    # Convert the tables to pandas DataFrames
    param_df = pd.DataFrame(param_rows, columns=param_headers)
    result_df = pd.DataFrame(result_rows, columns=result_headers)

    # Export to an Excel file
    if file_path == '':
        now = datetime.now()
        formatted_time = now.strftime("%Y%m%d%H%M%S")
        file_path = f'./output/output_tables_{epochs}_{learning_rate}_{formatted_time}.xlsx'

    # Load the existing workbook
    sheet_name = f's_{instance}'

    try:
        # Try to load the existing workbook
        workbook = load_workbook(file_path)
    except FileNotFoundError:
        # Create a new workbook if it doesn't exist
        workbook = Workbook()
        workbook.remove(workbook.active)  # Remove the default sheet

    # Create a new sheet for the instance
    if sheet_name in workbook.sheetnames:
        print(f"Sheet '{sheet_name}' already exists. Overwriting it.")
        workbook.remove(workbook[sheet_name])
    worksheet = workbook.create_sheet(title=sheet_name)

    # Write param_df to the sheet
    for r_idx, row in enumerate(dataframe_to_rows(param_df, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            worksheet.cell(row=r_idx, column=c_idx, value=str(value))

    # Write result_df below param_df, with a blank row in between
    start_row = len(param_df) + 3
    for r_idx, row in enumerate(dataframe_to_rows(result_df, index=False, header=True), start=start_row):
        for c_idx, value in enumerate(row, start=1):
            worksheet.cell(row=r_idx, column=c_idx, value=value)

    # Write confusion_matrix below result_df, with a blank row in between
    start_column = 1
    for cm_name, cm_df in cm_dict.items():
        print(cm_name)
        start_row = len(result_df) + len(param_df) + 9
        start_column = start_column
        worksheet.cell(row=start_row - 2, column=start_column + 2, value=cm_name)
        for r_idx, row in enumerate(dataframe_to_rows(cm_df, index=True, header=True), start=start_row):
            for c_idx, value in enumerate(row, start_column):
                worksheet.cell(row=r_idx, column=c_idx, value=value)
        start_column = start_column + 5

    # Write classification report below confusion matrix, with a blank row in between
    start_column = 1
    for report_name, report_df in report_dict.items():
        print(report_name)
        start_row = len(result_df) + len(param_df) + 19
        start_column = start_column
        worksheet.cell(row=start_row - 2, column=start_column + 2, value=report_name)
        for r_idx, row in enumerate(dataframe_to_rows(report_df, index=True, header=True), start=start_row):
            for c_idx, value in enumerate(row, start_column):
                worksheet.cell(row=r_idx, column=c_idx, value=value)
        start_column = start_column + 6
    # Save the workbook
    workbook.save(file_path)
    print(f"Tables saved to {instance} in {file_path}.")
    return file_path
#%% md
# ### Main Execution
#%%
def execute_network(features, labels, feature_type, results_dict, model_dict, learning_rate, epochs, cm_dictionary,
                    report_dictionary):
    print('{} : Executing Network for {} Features ****************************'.format(datetime.now(), feature_type))

    # Initialize the hybrid neural network
    hybrid_nn = HybridNeuralNetwork(input_shape=(62, 62, 4), encoding_dim=128, num_classes=3,
                                    learning_rate=learning_rate, epochs=epochs)
    if 'SC' in feature_type:
        hybrid_nn = HybridNeuralNetwork(input_shape=(62, 4, 4), encoding_dim=128, num_classes=3,
                                        learning_rate=learning_rate, epochs=epochs)

    # Build the model
    hybrid_nn.build_model()

    # Compile the model
    hybrid_nn.compile_model()

    # Split Data
    X_train_1, X_test, y_train_1, y_test = split_data(features, labels)

    X_train, X_val, y_train, y_val = split_data(X_train_1, y_train_1)

    # Train the model (replace X_train, y_train, X_val, y_val with actual data)
    history = hybrid_nn.train_model(X_train, y_train, X_val, y_val, batch_size=32)

    plot_history(history, 'Training and Validation ', f'accuracy-{feature_type}')
    plot_history(history, 'Training and Validation ', f'loss-{feature_type}')

    # Predict (replace X with actual data)
    # predictions = hybrid_nn.predict(X)

    # Saving models
    model_dict[f'hybrid_{feature_type}'] = hybrid_nn

    # Printing and saving training average results
    dictionary = print_history_average(history, results_dict, feature_type)

    # Evaluate the model (replace X_test, y_test with actual data)
    test_loss, test_accuracy = hybrid_nn.evaluate_model(X_test, y_test)
    dictionary[f'{feature_type}_test'] = {'loss': test_loss, 'accuracy': test_accuracy}

    # Predicting for confusion matrix
    y_pred_prob = hybrid_nn.predict(X_test)  # Get probabilities
    y_pred = np.argmax(y_pred_prob, axis=1)  # Convert probabilities to class indices
    cm = confusion_matrix(y_test, y_pred)

    print(f'Test : {np.unique(y_test)}')
    print(f'Pred : {np.unique(y_pred)}')

    # Handling error due to prediction having fewer classes
    y_test_unique = np.unique(y_test)
    y_pred_unique = np.unique(y_pred)
    if len(y_pred_unique) < len(y_test_unique):
        # Find the missing classes in y_pred_unique
        missing_classes = np.setdiff1d(y_test_unique, y_pred_unique)

        # Add the missing classes to y_pred_unique
        y_pred_unique = np.concatenate((y_pred_unique, missing_classes))

        # Sort y_pred_unique to maintain order (optional)
        y_pred_unique = np.sort(y_pred_unique)

    cm_df = pd.DataFrame(
        cm,
        index=[f"Actual_{cls}" for cls in y_test_unique],
        columns=[f"Predicted_{cls}" for cls in y_pred_unique]
    )

    # Generate classification report
    report = classification_report(y_test, y_pred, output_dict=True)
    report_df = pd.DataFrame(report).transpose()
    print(report_df)
    report_df.index.name = "Class"

    cm_dictionary[f'Confusion Matricx {feature_type}'] = cm_df
    report_dictionary[f'Classification Report {feature_type}'] = report_df

    print('{} : Executing Network for {} Features **************************** : DONE'.format(datetime.now(),
                                                                                              feature_type))

    return dictionary, model_dict, cm_dictionary, report_dictionary
#%% md
# ### Calling main
#%%
# Trying multiple learning rates
learning_rates = np.linspace(0.001, 0.001, num=1)
epochs = 1000
learning_rate = 0.0001
file_path = ''
for instance in range(0, 1):
    results_dictionary = {}
    model_dictionary = {}
    cm_dictionary = {}
    report_dictionary = {}
    results_dictionary, model_dictionary, cm_dictionary, report_dictionary = execute_network(pcc_matrix_8s, labels_8s,
                                                                                             'PCC_8s',
                                                                                             results_dictionary,
                                                                                             model_dictionary,
                                                                                             learning_rate, epochs,
                                                                                             cm_dictionary,
                                                                                             report_dictionary)
    results_dictionary, model_dictionary, cm_dictionary, report_dictionary = execute_network(pcc_matrix_12s, labels_12s,
                                                                                             'PCC_12s',
                                                                                             results_dictionary,
                                                                                             model_dictionary,
                                                                                             learning_rate, epochs,
                                                                                             cm_dictionary,
                                                                                             report_dictionary)

    results_dictionary, model_dictionary, cm_dictionary, report_dictionary = execute_network(conn_matrix_8s, labels_8s,
                                                                                             'CONN_8s',
                                                                                             results_dictionary,
                                                                                             model_dictionary,
                                                                                             learning_rate, epochs,
                                                                                             cm_dictionary,
                                                                                             report_dictionary)
    results_dictionary, model_dictionary, cm_dictionary, report_dictionary = execute_network(conn_matrix_12s,
                                                                                             labels_12s, 'CONN_12s',
                                                                                             results_dictionary,
                                                                                             model_dictionary,
                                                                                             learning_rate, epochs,
                                                                                             cm_dictionary,
                                                                                             report_dictionary)

    results_dictionary, model_dictionary, cm_dictionary, report_dictionary = execute_network(pca_matrix_8s, labels_8s,
                                                                                             'PCA_8s',
                                                                                             results_dictionary,
                                                                                             model_dictionary,
                                                                                             learning_rate, epochs,
                                                                                             cm_dictionary,
                                                                                             report_dictionary)
    results_dictionary, model_dictionary, cm_dictionary, report_dictionary = execute_network(pca_matrix_12s, labels_12s,
                                                                                             'PCA_12s',
                                                                                             results_dictionary,
                                                                                             model_dictionary,
                                                                                             learning_rate, epochs,
                                                                                             cm_dictionary,
                                                                                             report_dictionary)

    results_dictionary, model_dictionary, cm_dictionary, report_dictionary = execute_network(sc_matrix_8s, labels_8s,
                                                                                             'SC_8s',
                                                                                             results_dictionary,
                                                                                             model_dictionary,
                                                                                             learning_rate, epochs,
                                                                                             cm_dictionary,
                                                                                             report_dictionary)
    results_dictionary, model_dictionary, cm_dictionary, report_dictionary = execute_network(sc_matrix_12s, labels_12s,
                                                                                             'SC_12s',
                                                                                             results_dictionary,
                                                                                             model_dictionary,
                                                                                             learning_rate, epochs,
                                                                                             cm_dictionary,
                                                                                             report_dictionary)

    file_path = save_params_and_results(model_dictionary, results_dictionary, cm_dictionary, report_dictionary,
                                        learning_rate, epochs, file_path, instance=instance)